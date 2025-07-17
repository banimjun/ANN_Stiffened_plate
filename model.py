import numpy as np
import pandas as pd
import tensorflow as tf
from tensorflow import keras
from dataset import DatasetSelect
import matplotlib.pyplot as plt
from sklearn.preprocessing import StandardScaler, MinMaxScaler
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, Color
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from keras.callbacks import EarlyStopping
from keras.models import load_model
pd.set_option('display.float_format', '{:,.1f}'.format)

class ULSPredict:
    def __init__(self, args):
        self.data_columns = args.data_columns
        self.epochs = args.epochs
        self.batch_size = args.batch_size
        self.method = args.method
        dataset = DatasetSelect(args.data_dir, self.data_columns, args.method)
        self.x_train, self.x_test, self.y_train, self.y_test, self.raw_dataset = dataset()

    def model_compile(self):
        print(keras.__version__)
        dropout_rate = 0.01
        model = keras.models.Sequential([
            keras.layers.Dense(units=13, input_shape=(13,), activation='relu'),
            keras.layers.Dense(units=26, activation='relu'),
            keras.layers.Dense(units=52, activation='relu'),
            keras.layers.Dense(units=104, activation='relu'),
            keras.layers.Dense(units=208, activation='relu'),
            keras.layers.Dense(units=208, activation='relu'),
            keras.layers.Dense(units=104, activation='relu'),
            keras.layers.Dense(units=52, activation='relu'),
            keras.layers.Dense(units=26, activation='relu'),
            keras.layers.Dense(units=13, activation='relu'),
            keras.layers.Dense(units=1)])
        optimizer = tf.keras.optimizers.Adam()

        model.compile(optimizer=optimizer, loss='mean_squared_error', metrics=['mean_absolute_error'])
        early_stopping = keras.callbacks.EarlyStopping(monitor='val_loss', min_delta=0, patience=100, verbose=1)
        model.summary()

        history = model.fit(
            self.x_train[self.data_columns],
            self.y_train,
            callbacks=[early_stopping],
            epochs=self.epochs,
            batch_size=self.batch_size,
            verbose=1,
            validation_data=(
                self.x_test[self.data_columns],
                self.y_test))

        model.save('saved_model_ULS.h5')

        prediction = pd.DataFrame(
            model.predict(self.x_test[self.data_columns]),
            columns=['predicted value'])
        initial_result = pd.concat(
            [self.y_test.reset_index(drop=True), prediction],
            axis=1)

        predict_result_pd = prediction[['predicted value']]

        plt.plot(history.history['mean_absolute_error'])
        plt.plot(history.history['val_mean_absolute_error'])
        plt.title('model mae')
        plt.ylabel('mean_absolute_error')
        plt.xlabel('epoch')
        plt.legend(['train', 'test'], loc='upper right')
        plt.savefig('model_mae.png')

        return history, prediction, initial_result, predict_result_pd, self.x_train, self.x_test, self.y_train, self.y_test

    def inverse_norm_method(self, prediction):

        if self.method == 'std':
            self.inv_normalization_std(prediction)

        elif self.method == 'minmax':
            self.inv_normalization_minmax(prediction)

        return prediction

    def inv_normalization_std(self, prediction):

        standard = StandardScaler()
        standard_input_data = pd.DataFrame(standard.fit_transform(self.raw_dataset))
        inverse_x_test = pd.DataFrame(standard.inverse_transform(self.x_test), columns=self.data_columns)
        self.Excel_save(inverse_x_test,prediction)
        return inverse_x_test, prediction

    def inv_normalization_minmax(self, prediction):

        minmax = MinMaxScaler()
        standard_input_data = pd.DataFrame(minmax.fit_transform(self.raw_dataset))
        inverse_x_test = pd.DataFrame(minmax.inverse_transform(self.x_test), columns=self.data_columns)
        self.Excel_save(self.inverse_norm_method,self.prediction)
        return inverse_x_test, prediction

    def Excel_save(self,inverse_x_test, prediction):

        y_test = pd.DataFrame(self.y_test.reset_index(drop=True))

        predict_result_pd = prediction[['predicted value']]

        cal_error = abs(((y_test['uls'] - predict_result_pd['predicted value']) / y_test['uls']) * 100)
        cal_error = pd.DataFrame(cal_error, columns=['Error(%)'])
        cal_error_max = cal_error.max()
        cal_error_min = cal_error.min()
        cal_error_mean = cal_error.mean()

        excel_test_input = pd.concat([inverse_x_test, y_test, predict_result_pd, cal_error], axis=1)
        excel_test_input.to_excel('ULS_Predict.xlsx')

        excel_file = openpyxl.load_workbook('D:\\외장하드\\부산대학교\\2.박사과정\\2025\\BK논문투고\\Revise_2\Github\\ULS\\ULS_Predict.xlsx')
        ws = excel_file.active

        col_names = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        for col_list in col_names:
            ws.column_dimensions[str(col_list)].width = 7

        col_names = ['M', 'N', 'O']
        for col_list in col_names:
            ws.column_dimensions[str(col_list)].width = 20

        align_center = Alignment(horizontal='center', vertical='center')
        for row in ws['A1:O100']:
            for cell in row:
                cell.alignment = align_center

        ws['R1'] = "Summary"
        ws['R2'] = "MAX" + str(cal_error_max.values)
        ws['R3'] = "MIN" + str(cal_error_min.values)
        ws['R4'] = "Average" + str(cal_error_mean.values)

        print(cal_error_mean, cal_error_min, cal_error_max)
        excel_file.save('ULS_Predict.xlsx')
        return excel_file
